#pyinstaller --onefile --noconsole realtor_scrapper.py




import threading
import time
import sys
import traceback

import customtkinter as ctk

# --- Selenium imports ---
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException

# -------------------- Selenium Driver Setup --------------------




import json
import logging
import os
import pickle
import re
import subprocess
import sys
import time
from glob import glob
from typing import List, Optional

import undetected_chromedriver as uc
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    ElementClickInterceptedException,
    ElementNotInteractableException,
    WebDriverException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

from selenium.webdriver.common.keys import Keys
# =======================
# CONFIG & CONSTANTS
# =======================
BASEDIR = os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__))

LOG_FILE = os.path.join(BASEDIR, "scraper.log")


# =======================
# LOGGER SETUP
# =======================
logger = logging.getLogger("YELLOSCRAPPER")
logger.setLevel(logging.INFO)
if not logger.hasHandlers():
    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
    stream_handler = logging.StreamHandler()
    file_handler.setFormatter(formatter)
    stream_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)


def get_chrome_major_version() -> int:
    """
    Detects the installed Chrome major version.
    Works on Windows and Linux/Mac fallback.
    """
    try:
        # Windows registry query
        output = subprocess.check_output(
            r'reg query "HKEY_CURRENT_USER\Software\Google\Chrome\BLBeacon" /v version',
            shell=True, text=True
        )
        match = re.search(r"version\s+REG_SZ\s+([\d.]+)", output)
        if match:
            version = int(match.group(1).split('.')[0])
            logger.info(f"Chrome version detected from registry: {version}")
            return version
    except Exception:
        logger.debug("Windows registry query failed, trying fallback.")

    try:
        # Fallback: run 'chrome --version' (Linux/macOS)
        output = subprocess.check_output(["chrome", "--version"], text=True)
        match = re.search(r"(\d+)\.\d+\.\d+\.\d+", output)
        if match:
            version = int(match.group(1))
            logger.info(f"Chrome version detected from 'chrome --version': {version}")
            return version
    except Exception:
        logger.debug("Fallback chrome --version check failed.")

    logger.error("Could not detect Chrome version. Please ensure Chrome is installed and in PATH.")
    sys.exit(1)


def init_driver(headless: bool = False) -> uc.Chrome:
    """
    Initialize undetected_chromedriver with appropriate options.
    """
    version = get_chrome_major_version()
    logger.info(f"Initializing ChromeDriver with Chrome version {version} (headless={headless})")

    options = uc.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--no-sandbox")
    prefs = {
        "profile.default_content_setting_values.notifications": 2,  # Block notifications
        "profile.default_content_setting_values.geolocation": 2,    # Block location
        "profile.default_content_setting_values.media_stream_mic": 2,   # Block mic
        "profile.default_content_setting_values.media_stream_camera": 2, # Block camera
        "profile.default_content_setting_values.popups": 0           # Block popups
    }
    options.add_experimental_option("prefs", prefs)

    


    #options.add_argument("--blink-settings=imagesEnabled=false")
    

    driver = uc.Chrome(version_main=version, options=options)
    driver.maximize_window()
    return driver




import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter



def append_to_excel(data: dict, filename="scrapper.xlsx", sheet_name="Sheet1"):
    """
    Appends scraped data to Excel in a structured format.
    """

    headers = [
        "Price", "page link", "1st Image link",
        "Line 1 Address", "City", "Province", "Post Code",
        "Salesperson 1", "Phone#1", "Phone#2",
        "Brokerage1", "Brokerage1 Addr#" ,"Brokerage1 Tel#",
        "Salesperson 2", "Phone#1", "Phone#2",
        "Brokerage2", "Brokerage2 Addr#" ,"Brokerage2 Tel#"
    ]

    # Create workbook if not exists
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        wb.save(filename)

    # Load existing workbook
    wb = load_workbook(filename)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
    else:
        ws = wb[sheet_name]

    # ----- Parse data into required format -----
    # Address split
    line1, city, province, postal = "", "", "", ""
    if "address" in data and data["address"]:
        parts = data["address"].split("\n")
        if len(parts) >= 2:
            line1 = parts[0].strip()
            # Example: "Norwich (Norwich Town), Ontario N0J1P0"
            addr_parts = parts[1].split(",")
            if len(addr_parts) >= 2:
                city = addr_parts[0].strip()
                province_post = addr_parts[1].strip().split(" ")
                if len(province_post) >= 2:
                    province = province_post[0]
                    postal = " ".join(province_post[1:])

    # Salesperson 1 (main agent)
    salesperson1 = data.get("salesperson1", "")
    phone1 = data.get("salesperson1_phone1", "")
    phone2 = data.get("salesperson1_phone2", "")   



    salesperson2 = data.get("salesperson2", "")
    salesperson2_phone1 = data.get("salesperson2_phone1", "")
    salesperson2_phone2 = data.get("salesperson2_phone2", "")


    #brokara


    brokerage1 = data.get("brokerage1", "")
    brokerage1_address = data.get("brokerage1_address", "")
    brokerage1_tel = data.get("brokerage1_tel", "")

    brokerage2 = data.get("brokerage2", "")
    brokerage2_address = data.get("brokerage1_address", "")
    brokerage2_tel = data.get("brokerage1_tel", "")



    # Row in correct order
    row = [
        data.get("price", ""),
        data.get("url", ""),
        data.get("image", ""),
        line1, city, province, postal,
        salesperson1, phone1, phone2,
        brokerage1, brokerage1_address,brokerage1_tel,
        salesperson2, salesperson2_phone1, salesperson2_phone2,
        brokerage2, brokerage2_address,brokerage2_tel
        
    ]

    # Append row
    ws.append(row)

    # Auto-adjust column width
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(len(str(cell.value)) for cell in ws[col_letter])
        ws.column_dimensions[col_letter].width = max(15, min(max_length + 2, 60))

    wb.save(filename)







def _safe_text(el, default="-"):
    try:
        t = el.text.strip()
        return t if t else default
    except Exception:
        return default



def get_listing_info(driver, timeout=10):
    wait = WebDriverWait(driver, timeout)
    info = {
        "image": "",
        "price": "",
        "address": "",
        # Salesperson 1
        "salesperson1": "-",
        "salesperson1_phone1": "-",
        "salesperson1_phone2": "-",
        # Salesperson 2
        "salesperson2": "-",
        "salesperson2_phone1": "-",
        "salesperson2_phone2": "-",
        # Brokerage / office 1
        "brokerage1": "-",
        "brokerage1_address":"-",
        "brokerage1_tel": "-",
        # Brokerage / office 2
        "brokerage2": "-",
        "brokerage2_address":"-",
        "brokerage2_tel": "-",
        # optional
        "url": ""
    }

    # ---- Basic single-element fields ----
    try:
        img = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='heroImage']")))
        info["image"] = img.get_attribute("src") or ""
    except Exception:
        info["image"] = ""

    try:
        price = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='listingPriceValue']")))
        info["price"] = _safe_text(price, default="")
    except Exception:
        info["price"] = ""

    try:
        listingAddress = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='listingAddress']")))
        info["address"] = _safe_text(listingAddress, default="")
    except Exception:
        info["address"] = ""

    # If you want page URL (current tab)
    try:
        info["url"] = driver.current_url
    except Exception:
        info["url"] = ""

    # ---- Realtor cards (salespersons) ----
    try:
        realtor_cards = wait.until(
            EC.presence_of_all_elements_located(
                (By.XPATH, "//*[starts-with(@id,'realtorCard')]//div[contains(@class,'realtorCardCon card ')]")
            )
        )
    except Exception:
        realtor_cards = []

    # Extract up to 2 salespersons
    for idx in range(2):
        if idx < len(realtor_cards):
            card = realtor_cards[idx]
            # Name
            try:
                name_el = card.find_element(By.XPATH, ".//*[@class='realtorCardName']")
                name = _safe_text(name_el, default="-")
            except Exception:
                name = "-"
            # Telephones (may be multiple)
            try:
                phone_els = card.find_elements(By.XPATH, ".//*[@data-type='Telephone']")
                phones = [p.text.strip() for p in phone_els if p.text.strip()]
            except Exception:
                phones = []

            phone1 = phones[0] if len(phones) >= 1 else "-"
            phone2 = phones[1] if len(phones) >= 2 else "-"

            if idx == 0:
                info["salesperson1"] = name
                info["salesperson1_phone1"] = phone1
                info["salesperson1_phone2"] = phone2
            else:
                info["salesperson2"] = name
                info["salesperson2_phone1"] = phone1
                info["salesperson2_phone2"] = phone2
        else:
            # no card for this index -> keep defaults
            pass

    # ---- Office / brokerage cards ----
    try:
        office_cards = wait.until(
            EC.presence_of_all_elements_located((By.XPATH, "//*[starts-with(@id,'officeCard')]"))
        )
    except Exception:
        office_cards = []

    for idx in range(2):
        if idx < len(office_cards):
            card = office_cards[idx]

            # Office info text (contains brokerage name + 'Brokerage' + address lines)
            try:
                office_info_el = card.find_element(By.XPATH, ".//*[@class='officeCardTopLeft']")
                office_info_text = _safe_text(office_info_el, default="-")
            except Exception:
                office_info_text = _safe_text(card, default="-")

            # Split into lines
            lines = office_info_text.splitlines() if office_info_text and office_info_text != "-" else []

            # Brokerage name = first line
            brokerage_name = lines[0].strip() if len(lines) > 0 else "-"

            # Address = everything after the first 2 lines (skip brokerage name + "Brokerage")
            brokerage_address = " ".join(line.strip() for line in lines[2:]) if len(lines) > 2 else "-"

            # Office phone(s)
            phones = []
            try:
                tel_els = card.find_elements(By.XPATH, ".//*[@class='officeCardContactNumber']")
                if not tel_els:
                    tel_els = card.find_elements(By.XPATH, ".//*[@data-type='Telephone']")
                phones = [t.text.strip() for t in tel_els if t.text.strip()]
            except Exception:
                phones = []

            brokerage_tel = phones[0] if len(phones) >= 1 else "-"

            # Assign to info dict
            if idx == 0:
                info["brokerage1"] = brokerage_name
                info["brokerage1_address"] = brokerage_address
                info["brokerage1_tel"] = brokerage_tel
            else:
                info["brokerage2"] = brokerage_name
                info["brokerage2_address"] = brokerage_address
                info["brokerage2_tel"] = brokerage_tel


    return info






def process(driver):
    try:
        items = driver.find_elements(By.XPATH, "//*[@data-binding='href=DetailsURL']")
    except Exception:
        print("Cannot load the main page...")
        while True:
            input("refresh ?")
            driver.refresh()
            time.sleep(2)
            process(driver)  

    while items == []:
        print("Cannot load the main page...")
        while True:
            input("refresh ?")
            driver.refresh()
            time.sleep(2)
            process(driver) 

    print(f"Total item {len(items)} found")

    for idx, eachitem in enumerate(items):
        print(f"{idx+1} / {len(items)} running")        

        # Scroll into view before clicking
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", eachitem)
        time.sleep(0.5)

        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(eachitem))
            ActionChains(driver).move_to_element(eachitem).click().perform()
        except Exception as e:
            print(f"⚠️ Click failed, trying JS click: {e}")
            driver.execute_script("arguments[0].click();", eachitem)

        time.sleep(1)
        driver.switch_to.window(driver.window_handles[-1])
        time.sleep(1)

        try:
            info = get_listing_info(driver, timeout=10)
            append_to_excel(info)
        except Exception as e:
            print(f"❌ cannot visit the item page {e}")
        finally:
            print("="*8)

        driver.close()
        driver.switch_to.window(driver.window_handles[-1])
        time.sleep(1)








def startbrowser(url):
    """Create driver, open url, return driver."""
    driver = init_driver()
    driver.get(url)
    return driver


# ---------------- Pagination Logic ----------------
def pagination(driver, log, stop_event):
    try:
        total = driver.find_element(By.ID, "mapResultsNumVal").text
        log(f"total item {total}")
    except Exception as e:
        log(f"[warn] Could not read total mapResultsNumVal: {e}")

    pagecount = 1
    while not stop_event.is_set():
        log(f"Clicked Next page  {pagecount}")
        try:
            # replace with your actual scraping logic
            time.sleep(3)

            process(driver)


            time.sleep(3)

            wait = WebDriverWait(driver, 15)
            next_btn = wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "paginationLinkForward"))
            )

            aria_label = next_btn.get_attribute("aria-label") or ""
            if "disabled" in aria_label.lower():
                log("Next button is disabled. Stopping.")
                break

            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_btn)
            next_btn.click()
            log("Clicked Next Page")
            pagecount += 1
            time.sleep(5)

        except TimeoutException:
            log("No more Next button found (timeout). Stopping.")
            break
        except WebDriverException as e:
            log(f"[webdriver] {e}")
            break
        except Exception as e:
            log(f"[error] {e}\n{traceback.format_exc()}")
            break

    log("Pagination loop finished.")






# ---------------- UI ----------------
class App(ctk.CTk):
    def __init__(self, driver):
        super().__init__()

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")

        self.title("Realtor.ca Pagination Controller")
        self.geometry("880x560")

        # State
        self.driver = driver
        self.worker = None
        self.stop_event = threading.Event()
        self._lock = threading.Lock()

        # Layout
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)

        # Header
        self.header = ctk.CTkFrame(self, corner_radius=16)
        self.header.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 8))
        self.header.grid_columnconfigure(1, weight=1)

        self.icon_label = ctk.CTkLabel(self.header, text="🔎", font=ctk.CTkFont(size=24))
        self.icon_label.grid(row=0, column=0, padx=(12, 8), pady=12)
        self.title_label = ctk.CTkLabel(self.header, text="Realtor Pagination", font=ctk.CTkFont(size=22, weight="bold"))
        self.title_label.grid(row=0, column=1, sticky="w", padx=(0, 8), pady=12)
        self.status_dot = ctk.CTkLabel(self.header, text="● idle", text_color="#9ca3af")
        self.status_dot.grid(row=0, column=2, padx=12, pady=12)

        # URL Row
        self.url_frame = ctk.CTkFrame(self, corner_radius=16)
        self.url_frame.grid(row=1, column=0, sticky="ew", padx=16, pady=8)
        self.url_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(self.url_frame, text="URL").grid(row=0, column=0, padx=12, pady=12)
        self.url_entry = ctk.CTkEntry(self.url_frame, placeholder_text="Enter target URL...", height=40)
        self.url_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=12)

        default_url = "https://www.realtor.ca/map#..."
        self.url_entry.insert(0, default_url)

        self.open_btn = ctk.CTkButton(self.url_frame, text="Open URL", height=40, command=self.open_url)
        self.open_btn.grid(row=0, column=2, padx=8, pady=12)

        # Controls
        self.controls = ctk.CTkFrame(self, corner_radius=16)
        self.controls.grid(row=2, column=0, sticky="ew", padx=16, pady=8)
        self.controls.grid_columnconfigure((0, 1, 2, 3), weight=1)

        self.start_btn = ctk.CTkButton(self.controls, text="▶ Start Pagination", command=self.start_worker)
        self.start_btn.grid(row=0, column=0, padx=8, pady=12, sticky="ew")

        self.stop_btn = ctk.CTkButton(self.controls, text="■ Stop", command=self.stop_worker,
                                      fg_color="#ef4444", hover_color="#b91c1c")
        self.stop_btn.grid(row=0, column=1, padx=8, pady=12, sticky="ew")

        self.quit_btn = ctk.CTkButton(self.controls, text="✕ Quit", command=self.safe_quit)
        self.quit_btn.grid(row=0, column=3, padx=8, pady=12, sticky="ew")

        # Log
        self.log_box = ctk.CTkTextbox(self, height=380)
        self.log_box.grid(row=3, column=0, sticky="nsew", padx=16, pady=16)

    # ---------- Helpers ----------
    def set_status(self, text, color="#9ca3af"):
        self.status_dot.configure(text=f"● {text}", text_color=color)

    def log(self, message: str):
        with self._lock:
            self.log_box.insert("end", f"{message}\n")
            self.log_box.see("end")
        print(message)

    def open_url(self):
        url = self.url_entry.get().strip()
        if not url:
            self.log("[warn] URL is empty.")
            return
        try:
            self.driver.get(url)   # ✅ reuse existing driver
            self.log(f"Opened: {url}")
        except Exception as e:
            self.log(f"[error] Failed to open URL: {e}")
            self.set_status("driver error", "#ef4444")

    def start_worker(self):
        if self.worker and self.worker.is_alive():
            self.log("[info] Worker already running.")
            return
        if not self.driver:
            self.log("[error] Driver not available.")
            return

        self.stop_event.clear()
        self.set_status("running", "#22c55e")
        self.worker = threading.Thread(target=self._run_pagination, daemon=True)
        self.worker.start()

    def _run_pagination(self):
        try:
            pagination(self.driver, self.log, self.stop_event)
        except Exception as e:
            self.log(f"[fatal] {e}\n{traceback.format_exc()}")
        finally:
            self.set_status("idle", "#9ca3af")

    def stop_worker(self):
        self.stop_event.set()
        self.set_status("stopping", "#f59e0b")
        self.log("Stop requested. Waiting for loop to finish...")

    def safe_quit(self):
        try:
            self.stop_event.set()
            if self.worker and self.worker.is_alive():
                self.worker.join(timeout=5)
        except Exception:
            pass
        try:
            if self.driver:
                self.driver.quit()
                self.log("Driver closed.")
        except Exception:
            pass
        self.destroy()


# ---------------- Main ----------------
def main():
    default_url = "https://www.realtor.ca/map#ZoomLevel=9&Center=42.949006%2C-81.248535&LatitudeMax=43.25883&LongitudeMax=-79.99335&LatitudeMin=42.63762&LongitudeMin=-82.50372&Sort=6-D&PGeoIds=g30_dpwhr7kj&GeoName=London%2C%20ON&PropertyTypeGroupID=1&TransactionTypeId=2&PropertySearchTypeId=0&Currency=CAD"
    driver = startbrowser(default_url)   # ✅ open browser immediately

    app = App(driver)
    app.log(f"Opened on startup: {default_url}")
    app.mainloop()


if __name__ == "__main__":
    main()

